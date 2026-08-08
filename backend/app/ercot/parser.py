"""Parse a GIS Report .xlsx into normalized Project records.

The workbook has four project-bearing sheets. Each has a title/notes preamble
followed by a header row whose first cell is "INR", then data rows. We locate
the header dynamically (row counts differ month to month) and map columns by
normalized header name, so column reordering doesn't break parsing.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

import openpyxl

from ..models import Project, ProjectStatus, SizeCategory
from . import codes

# sheet name -> (status, size_category override or None -> read from row)
_SHEETS: dict[str, tuple[ProjectStatus, SizeCategory | None]] = {
    "Project Details - Large Gen": ("active", "Large"),
    "Project Details - Small Gen": ("active", "Small"),
    "Inactive Projects": ("inactive", None),
    "Cancellation Update": ("cancelled", None),
}


def _norm(header: Any) -> str:
    return " ".join(str(header).strip().lower().split()) if header is not None else ""


def _find_header_row(rows: list[tuple]) -> int | None:
    for i, row in enumerate(rows):
        if row and _norm(row[0]) == "inr":
            return i
    return None


def _to_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _to_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    return None


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


# Column lookups tolerate the small header differences between sheets.
def _get(row_map: dict[str, Any], *names: str) -> Any:
    for n in names:
        if n in row_map and row_map[n] not in (None, ""):
            return row_map[n]
    return None


def parse_workbook(
    content: bytes,
    *,
    report_date: date | None = None,
    source_file: str | None = None,
    technologies: set[str] | None = codes.DATA_CENTER_TECHNOLOGIES,
) -> list[Project]:
    """Parse the GIS Report into Project records.

    `technologies` restricts results to those technology codes (upper-cased).
    Defaults to the data-center-relevant set (CC, GT, IC, BA, EN, FC). Pass
    ``None`` to keep every project regardless of technology.

    Sheets that carry a Technology column (Large/Small Gen) are filtered by that
    set. The Inactive / Cancellation sheets have no Technology column, so they
    are filtered by the equivalent fuels (Gas, Battery Storage) instead, keeping
    them on the same thesis.
    """
    allowed = {t.upper() for t in technologies} if technologies else None
    allowed_fuels = codes.DATA_CENTER_FUELS if allowed is not None else None
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    projects: list[Project] = []

    for sheet_name, (status, size_override) in _SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        hidx = _find_header_row(rows)
        if hidx is None:
            continue
        headers = [_norm(h) for h in rows[hidx]]
        has_tech_column = "technology" in headers

        for raw_row in rows[hidx + 1 :]:
            if not raw_row or raw_row[0] in (None, ""):
                continue
            inr = str(raw_row[0]).strip()
            if not inr:
                continue

            row_map = {
                headers[j]: raw_row[j]
                for j in range(min(len(headers), len(raw_row)))
                if headers[j]
            }
            raw_json = {k: _jsonable(v) for k, v in row_map.items()}

            fuel = _get(row_map, "fuel")
            fuel = str(fuel).strip() if fuel is not None else None
            tech = _get(row_map, "technology")
            tech = str(tech).strip() if tech is not None else None

            if allowed is not None:
                if has_tech_column:
                    if (tech or "").upper() not in allowed:
                        continue
                # Status sheets (no Technology column) fall back to fuel.
                elif (fuel or "").upper() not in allowed_fuels:
                    continue

            size = size_override
            if size is None:
                sc = _get(row_map, "size category")
                if sc:
                    size = "Large" if "large" in str(sc).lower() else "Small"

            projects.append(
                Project(
                    inr=inr,
                    project_name=_get(row_map, "project name"),
                    size_category=size,
                    status=status,
                    interconnecting_entity=_get(row_map, "interconnecting entity"),
                    poi_location=_get(row_map, "poi location"),
                    county=_get(row_map, "county"),
                    cdr_reporting_zone=_get(row_map, "cdr reporting zone"),
                    fuel=fuel,
                    fuel_label=codes.fuel_label(fuel),
                    technology=tech,
                    technology_label=codes.technology_label(tech),
                    capacity_mw=_to_float(_get(row_map, "capacity (mw)", "mw **", "mw")),
                    gim_study_phase=_get(row_map, "gim study phase"),
                    projected_cod=_to_date(_get(row_map, "projected cod")),
                    ia_signed=(
                        str(_get(row_map, "ia signed")).strip()
                        if _get(row_map, "ia signed") is not None
                        else None
                    ),
                    approved_for_energization=_to_date(
                        _get(row_map, "approved for energization")
                    ),
                    approved_for_synchronization=_to_date(
                        _get(row_map, "approved for synchronization")
                    ),
                    inactive_date=_to_datetime(_get(row_map, "inactive date")),
                    cancel_date=_to_datetime(_get(row_map, "cancel date")),
                    comment=_get(row_map, "comment"),
                    report_date=report_date,
                    source_sheet=sheet_name,
                    source_file=source_file,
                    raw=raw_json,
                )
            )

    wb.close()
    return projects
